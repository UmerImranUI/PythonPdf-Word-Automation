import PyPDF2
import os
import re

import openpyxl


your_excel=openpyxl.load_workbook('a.xlsx')
your_sheet=your_excel['sheet1']

for file_name in os.listdir('all_format_pdf'):   #loop on files
    print(file_name)
    load_pdf=open('path')
    read_pdf=PyPDF2.PdfFileReader(load_pdf)   #load entire pdf in var
    page_count=read_pdf.getNumPages()
    first_page=read_pdf.getPage(0)            #read only the first page
    page_content=first_page.extractText()     #extract string output
    page_content=page_content.replace('\n','')#Replace New Line(if reqd)
    print(page_content)


    #10 dig mob Num extract
    mob_num=re.search(r'(?:\+?\d{2}[ -]?)?\d{10}', page_content).group()
    print(mob_num)

    #Email ID extract
    email_id=re.search(r'([a-zA-Z0-9._-]+@[a-zA-Z0-9._-]+\.[a-zA-z0-9_-]+)', page_content).group()
    print(email_id)

    #Zip Code Ext
    zipcode=re.search(r'(?:\+?\d{2}[ -]?)?\d{6}', page_content).group()
    print(zipcode)


reg_1= '(?<=Name :)(.*)(?=Address)'
reg_2='(?<=Dear)(.*)(?=Area Zip)'
reg_3='(?<=Dear)(.*)(?=Email id)'

found_name=re.search("|".join([reg_1,reg_2,reg_3]), page_content).group()
print(found_name)

last_row_num=your_sheet.max_row
print(last_row_num)

your_sheet.cell(column=1, row=last_row_num+1).value=found_name
your_sheet.cell(column=2, row=last_row_num+1).value=mob_num
your_sheet.cell(column=3, row=last_row_num+1).value=email_id
your_sheet.cell(column=4, row=last_row_num+1).value=zipcode

your_excel.save('output.xlsx')